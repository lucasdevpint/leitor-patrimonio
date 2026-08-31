# api/routes/importacao.py

import io
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from api.deps import get_conn, requer_admin


router = APIRouter(prefix="/importacao", tags=["Importação"])


def _texto(valor):
    if valor is None:
        return None

    valor = str(valor).strip()

    if not valor:
        return None

    return valor


def _data(valor):
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")

    texto = str(valor).strip()

    if not texto:
        return None

    # Data no formato DD/MM/YYYY
    try:
        return datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        pass

    return texto


def _valor(valor):
    if valor is None or valor == "":
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    # Ex.: 10.415,00
    texto = texto.replace(".", "").replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return None


def _encontrar_cabecalho(ws):
    """
    Procura automaticamente a linha que contém
    REGISTRO PATRIMONIAL.
    """

    for numero_linha, row in enumerate(
        ws.iter_rows(values_only=True), 1
    ):
        for valor in row:
            if valor is None:
                continue

            texto = str(valor).strip().upper()

            if texto == "REGISTRO PATRIMONIAL":
                return numero_linha

    return None


def _mapear_colunas(ws, linha_cabecalho):
    """
    Converte os nomes da planilha em índices de coluna.
    """

    mapa = {}

    for indice, valor in enumerate(
        next(
            ws.iter_rows(
                min_row=linha_cabecalho,
                max_row=linha_cabecalho,
                values_only=True
            )
        ),
        1
    ):
        if valor is None:
            continue

        nome = str(valor).strip().upper()

        mapa[nome] = indice

    return mapa


def _valor_coluna(row, mapa, nome):
    indice = mapa.get(nome)

    if not indice:
        return None

    if indice > len(row):
        return None

    return row[indice - 1]
def _encontrar_localizacao(ws):
    """
    Procura a localização informada no cabeçalho da planilha.

    Exemplo:
    LOCALIZAÇÃO: RUA WASHINGTON LUIZ, 675, EDIF 4, PORTO ALEGRE, RS
    """

    for row in ws.iter_rows(values_only=True):
        for valor in row:
            if valor is None:
                continue

            texto = str(valor).strip()

            if texto.upper().startswith("LOCALIZAÇÃO:"):
                return _texto(
                    texto.split(":", 1)[1]
                )

    return None

def _normalizar_status(situacao):
    """
    Converte a situação oficial para os status usados
    pelo sistema.

    Quando a UERGS não informa situação, deixamos
    como Disponível por enquanto.
    """

    if not situacao:
        return "Disponível"

    texto = situacao.strip().lower()

    if "descarte" in texto:
        return "Descarte"

    if "manuten" in texto:
        return "Em manutenção"

    if "defeito" in texto:
        return "Defeito"

    if "dispon" in texto:
        return "Disponível"

    if "uso" in texto:
        return "Em uso"

    # Não inventamos uma situação.
    # O valor original continuará preservado.
    return "Disponível"


@router.post("/preview")
async def preview_importacao(
    arquivo: UploadFile = File(...),
    conn=Depends(get_conn),
    _=Depends(requer_admin)
):
    """
    Apenas analisa a planilha.

    NÃO insere, altera ou exclui registros.
    """

    if not arquivo.filename:
        raise HTTPException(
            status_code=400,
            detail="Nenhum arquivo foi enviado."
        )

    if not arquivo.filename.lower().endswith(
        (".xlsx", ".xlsm")
    ):
        raise HTTPException(
            status_code=400,
            detail="Envie uma planilha Excel (.xlsx ou .xlsm)."
        )

    conteudo = await arquivo.read()

    if not conteudo:
        raise HTTPException(
            status_code=400,
            detail="A planilha está vazia."
        )

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(conteudo),
            read_only=True,
            data_only=True
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Não foi possível ler a planilha: {e}"
        )

    if "Oficial" not in wb.sheetnames:
        wb.close()

        raise HTTPException(
            status_code=400,
            detail=(
                "A planilha não possui a aba 'Oficial'. "
                f"Abas encontradas: {', '.join(wb.sheetnames)}"
            )
        )

    ws = wb["Oficial"]

    localizacao = _encontrar_localizacao(ws)
    linha_cabecalho = _encontrar_cabecalho(ws)

    if not linha_cabecalho:
        wb.close()

        raise HTTPException(
            status_code=400,
            detail=(
                "Não encontrei o cabeçalho "
                "'REGISTRO PATRIMONIAL' na aba Oficial."
            )
        )

    mapa = _mapear_colunas(ws, linha_cabecalho)

    obrigatorias = [
        "REGISTRO PATRIMONIAL",
        "REGISTRO PATRIMONIAL ANTERIOR",
        "DT. AQUISIÇÃO",
        "DESCRIÇÃO DO BEM",
        "Nº DE SÉRIE",
        "SITUAÇÃO",
        "ÚLTIMO INVENTÁRIO",
        "VALOR UNIT. (R$)"
    ]

    faltando = [
        campo for campo in obrigatorias
        if campo not in mapa
    ]

    if faltando:
        wb.close()

        raise HTTPException(
            status_code=400,
            detail={
                "mensagem": "Colunas obrigatórias não encontradas.",
                "faltando": faltando,
                "colunas_encontradas": list(mapa.keys())
            }
        )

    # Busca todos os códigos já existentes.
    cursor = conn.cursor()
    cursor.execute("SELECT codigo FROM patrimonio")

    codigos_existentes = {
        str(row[0]).strip()
        for row in cursor.fetchall()
        if row[0] is not None
    }

    cursor.close()

    registros = []

    novos = 0
    existentes = 0
    sem_codigo = 0
    linhas_vazias = 0

    for numero_linha, row in enumerate(
        ws.iter_rows(
            min_row=linha_cabecalho + 1,
            values_only=True
        ),
        linha_cabecalho + 1
    ):

        codigo = _texto(
            _valor_coluna(
                row,
                mapa,
                "REGISTRO PATRIMONIAL"
            )
        )

        descricao = _texto(
            _valor_coluna(
                row,
                mapa,
                "DESCRIÇÃO DO BEM"
            )
        )

        # Ignora linhas totalmente vazias.
        if not codigo and not descricao:
            linhas_vazias += 1
            continue

        if not codigo:
            sem_codigo += 1

            registros.append({
                "linha": numero_linha,
                "localizacao":localizacao,
                "codigo": None,
                "codigo_secundario": _texto(
                    _valor_coluna(
                        row,
                        mapa,
                        "REGISTRO PATRIMONIAL ANTERIOR"
                    )
                ),
                "descricao": descricao,
                "numero_serie": _texto(
                    _valor_coluna(
                        row,
                        mapa,
                        "Nº DE SÉRIE"
                    )
                ),
                "status": _normalizar_status(
                    _texto(
                        _valor_coluna(
                            row,
                            mapa,
                            "SITUAÇÃO"
                        )
                    )
                ),
                "situacao_original": _texto(
                    _valor_coluna(
                        row,
                        mapa,
                        "SITUAÇÃO"
                    )
                ),
                "data_aquisicao": _data(
                    _valor_coluna(
                        row,
                        mapa,
                        "DT. AQUISIÇÃO"
                    )
                ),
                "registro_inventario": _texto(
                    _valor_coluna(
                        row,
                        mapa,
                        "ÚLTIMO INVENTÁRIO"
                    )
                ),
                "valor": _valor(
                    _valor_coluna(
                        row,
                        mapa,
                        "VALOR UNIT. (R$)"
                    )
                ),
                "situacao": "SEM_CODIGO"
            })

            continue

        if codigo in codigos_existentes:
            situacao = "EXISTENTE"
            existentes += 1
        else:
            situacao = "NOVO"
            novos += 1

        situacao_original = _texto(
            _valor_coluna(
                row,
                mapa,
                "SITUAÇÃO"
            )
        )

        registros.append({
            "linha": numero_linha,
            "localizacao":localizacao,
            "codigo": codigo,
            "codigo_secundario": _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "REGISTRO PATRIMONIAL ANTERIOR"
                )
            ),
            "descricao": descricao,
            "numero_serie": _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "Nº DE SÉRIE"
                )
            ),
            "status": _normalizar_status(
                situacao_original
            ),
            "situacao_original": situacao_original,
            "data_aquisicao": _data(
                _valor_coluna(
                    row,
                    mapa,
                    "DT. AQUISIÇÃO"
                )
            ),
            "registro_inventario": _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "ÚLTIMO INVENTÁRIO"
                )
            ),
            "valor": _valor(
                _valor_coluna(
                    row,
                    mapa,
                    "VALOR UNIT. (R$)"
                )
            ),
            "situacao": situacao
        })

    wb.close()

    return {
        "arquivo": arquivo.filename,
        "aba": "Oficial",
        "linha_cabecalho": linha_cabecalho,
        "resumo": {
            "total": len(registros),
            "novos": novos,
            "existentes": existentes,
            "sem_codigo": sem_codigo,
            "linhas_vazias": linhas_vazias
        },
        "registros": registros
    }
@router.post("/confirmar")
async def confirmar_importacao(
    arquivo: UploadFile = File(...),
    conn=Depends(get_conn),
    _=Depends(requer_admin)
):
    """
    Importa definitivamente os patrimônios da planilha.

    Apenas registros cujo código ainda não existe
    no banco serão inseridos.
    """

    if not arquivo.filename:
        raise HTTPException(
            status_code=400,
            detail="Nenhum arquivo foi enviado."
        )

    if not arquivo.filename.lower().endswith(
        (".xlsx", ".xlsm")
    ):
        raise HTTPException(
            status_code=400,
            detail="Envie uma planilha Excel (.xlsx ou .xlsm)."
        )

    conteudo = await arquivo.read()

    if not conteudo:
        raise HTTPException(
            status_code=400,
            detail="A planilha está vazia."
        )

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(conteudo),
            read_only=True,
            data_only=True
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Não foi possível ler a planilha: {e}"
        )

    if "Oficial" not in wb.sheetnames:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail="A planilha não possui a aba 'Oficial'."
        )

    ws = wb["Oficial"]

    linha_cabecalho = _encontrar_cabecalho(ws)

    if not linha_cabecalho:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail=(
                "Não encontrei o cabeçalho "
                "'REGISTRO PATRIMONIAL'."
            )
        )

    mapa = _mapear_colunas(ws, linha_cabecalho)

    obrigatorias = [
        "REGISTRO PATRIMONIAL",
        "REGISTRO PATRIMONIAL ANTERIOR",
        "DT. AQUISIÇÃO",
        "DESCRIÇÃO DO BEM",
        "Nº DE SÉRIE",
        "SITUAÇÃO",
        "ÚLTIMO INVENTÁRIO",
        "VALOR UNIT. (R$)"
    ]

    faltando = [
        campo for campo in obrigatorias
        if campo not in mapa
    ]

    if faltando:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail={
                "mensagem": "Colunas obrigatórias não encontradas.",
                "faltando": faltando
            }
        )

    cursor = conn.cursor()

    try:
        # Busca novamente os códigos existentes.
        # Isso é proposital: evita duplicação mesmo que
        # alguém tenha inserido registros depois do preview.
        cursor.execute("SELECT codigo FROM patrimonio")

        codigos_existentes = {
            str(row[0]).strip()
            for row in cursor.fetchall()
            if row[0] is not None
        }

        importados = 0
        ignorados = 0
        sem_codigo = 0
        erros = []

        for numero_linha, row in enumerate(
            ws.iter_rows(
                min_row=linha_cabecalho + 1,
                values_only=True
            ),
            linha_cabecalho + 1
        ):

            codigo = _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "REGISTRO PATRIMONIAL"
                )
            )

            descricao = _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "DESCRIÇÃO DO BEM"
                )
            )

            # Ignora linhas vazias.
            if not codigo and not descricao:
                continue

            # Não permite patrimônio sem código.
            if not codigo:
                sem_codigo += 1
                continue

            # Ignora linhas de rodapé da planilha oficial.
            # Ex.: "QUANTIDADE DE BENS: 2163"
            if not codigo.isdigit():
                continue

            # Já existe: não duplica.
            if codigo in codigos_existentes:
                ignorados += 1
                continue

            situacao_original = _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "SITUAÇÃO"
                )
            )

            status = _normalizar_status(
                situacao_original
            )

            codigo_secundario = _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "REGISTRO PATRIMONIAL ANTERIOR"
                )
            )

            data_aquisicao = _data(
                _valor_coluna(
                    row,
                    mapa,
                    "DT. AQUISIÇÃO"
                )
            )

            numero_serie = _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "Nº DE SÉRIE"
                )
            )

            registro_inventario = _texto(
                _valor_coluna(
                    row,
                    mapa,
                    "ÚLTIMO INVENTÁRIO"
                )
            )

            valor = _valor(
                _valor_coluna(
                    row,
                    mapa,
                    "VALOR UNIT. (R$)"
                )
            )
            # Localização da unidade na planilha
            localizacao = None

            for linha_anterior in ws.iter_rows(
                min_row=1,
                max_row=linha_cabecalho - 1,
                values_only=True
            ):
                for valor_celula in linha_anterior:
                    if valor_celula is None:
                        continue

                    texto_local = str(valor_celula).strip()

                    if texto_local.upper().startswith("LOCALIZAÇÃO:"):
                        localizacao = texto_local.split(":", 1)[1].strip()
                        break

                if localizacao:
                    break




            try:
                cursor.execute(
                    """
                    INSERT INTO patrimonio (
                        codigo,
                        descricao,
                        status,
                        codigo_secundario,
                        data_aquisicao,
                        valor,
                        numero_serie,
                        registro_inventario,
                        situacao_original,
                        origem_importacao,
                        localizacao
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s
                    )
                    """,
                    (
                        codigo,
                        descricao or "Sem descrição",
                        status,
                        codigo_secundario,
                        data_aquisicao,
                        valor,
                        numero_serie,
                        registro_inventario,
                        situacao_original,
                        arquivo.filename,
                        localizacao
                    )
                )

                # Adiciona imediatamente ao conjunto.
                codigos_existentes.add(codigo)
                importados += 1

            except Exception as e:
                erros.append({
                    "linha": numero_linha,
                    "codigo": codigo,
                    "erro": str(e)
                })

        if erros:
            conn.rollback()
        else:
            conn.commit()

    except Exception as e:
        conn.rollback()
        cursor.close()
        wb.close()

        raise HTTPException(
            status_code=500,
            detail=f"Erro durante a importação: {e}"
        )

    cursor.close()
    wb.close()

    return {
        "mensagem": (
            "Importação concluída com sucesso."
            if not erros
            else "Importação cancelada devido a erros."
        ),
        "arquivo": arquivo.filename,
        "importados": importados,
        "ignorados_existentes": ignorados,
        "sem_codigo": sem_codigo,
        "erros": erros
    }
