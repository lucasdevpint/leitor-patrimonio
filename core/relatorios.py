# core/relatorios.py
# Geração de relatórios em Excel com múltiplas abas.

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False


PASTA_RELATORIOS = os.path.join(os.path.dirname(__file__), "..", "assets", "relatorios")
os.makedirs(PASTA_RELATORIOS, exist_ok=True)

# Paleta
COR_HEADER  = "1A3A5C"
COR_LINHA1  = "E8F0FE"
COR_LINHA2  = "FFFFFF"
COR_ALERTA  = "FFCCCC"
COR_OK      = "CCFFCC"


def _estilo_header(ws, linha: int, colunas: list):
    fill   = PatternFill("solid", fgColor=COR_HEADER)
    fonte  = Font(bold=True, color="FFFFFF", size=11)
    borda  = Border(bottom=Side(style="thin", color="AAAAAA"))
    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=linha, column=col, value=titulo)
        cell.fill   = fill
        cell.font   = fonte
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _estilo_linha(ws, linha: int, valores: list, alerta: bool = False):
    cor  = COR_ALERTA if alerta else (COR_LINHA1 if linha % 2 == 0 else COR_LINHA2)
    fill = PatternFill("solid", fgColor=cor)
    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=linha, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def gerar_relatorio_completo(conn):
    """
    Gera um arquivo Excel com abas:
    - Todos os Patrimônios
    - Por Status
    - Manutenções Pendentes
    - Log de Auditoria (últimos 500)
    Retorna o caminho do arquivo gerado.
    """
    if not EXCEL_OK:
        return None, "openpyxl não instalado. Execute: pip install openpyxl"

    cursor = conn.cursor(dictionary=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove aba padrão

    # ── Aba 1: Todos os Patrimônios ──────────────────────────
    ws1 = wb.create_sheet("Patrimônios")
    ws1.row_dimensions[1].height = 24
    colunas1 = ["Código", "Cód. Secundário", "Tipo", "Descrição",
                 "Marca", "Modelo", "Nº Série", "Status",
                 "Responsável", "Local", "Data Aquisição", "Valor (R$)"]
    _estilo_header(ws1, 1, colunas1)

    cursor.execute("""
        SELECT p.codigo, p.codigo_secundario, p.tipo_patrimonio, p.descricao,
               p.marca, p.modelo, p.numero_serie, p.status,
               p.responsavel, l.nome AS local,
               p.data_aquisicao, p.valor
        FROM patrimonio p
        LEFT JOIN locais l ON p.local_id = l.id
        ORDER BY p.codigo
    """)
    for i, row in enumerate(cursor.fetchall(), 2):
        vals = [
            row["codigo"], row["codigo_secundario"], row["tipo_patrimonio"],
            row["descricao"], row["marca"], row["modelo"], row["numero_serie"],
            row["status"], row["responsavel"], row["local"],
            str(row["data_aquisicao"]) if row["data_aquisicao"] else "",
            float(row["valor"]) if row["valor"] else "",
        ]
        _estilo_linha(ws1, i, vals, alerta=(row["status"] in ("Defeito", "Descarte")))
    _autofit(ws1)

    # ── Aba 2: Resumo por Status ─────────────────────────────
    ws2 = wb.create_sheet("Resumo por Status")
    _estilo_header(ws2, 1, ["Status", "Quantidade", "% do Total"])
    cursor.execute("""
        SELECT status, COUNT(*) AS qtd FROM patrimonio GROUP BY status ORDER BY qtd DESC
    """)
    rows_status = cursor.fetchall()
    total = sum(r["qtd"] for r in rows_status)
    for i, row in enumerate(rows_status, 2):
        pct = f"{row['qtd']/total*100:.1f}%" if total else "0%"
        _estilo_linha(ws2, i, [row["status"], row["qtd"], pct])
    _autofit(ws2)

    # ── Aba 3: Manutenções Pendentes ─────────────────────────
    ws3 = wb.create_sheet("Manutenções Pendentes")
    _estilo_header(ws3, 1, ["Código", "Patrimônio", "Local", "Tipo",
                             "Descrição", "Data Prevista", "Responsável", "Dias Restantes"])
    cursor.execute("""
        SELECT p.codigo, p.descricao AS patrimonio, l.nome AS local,
               m.tipo, m.descricao, m.data_prevista, m.responsavel,
               DATEDIFF(m.data_prevista, CURDATE()) AS dias
        FROM manutencoes m
        JOIN patrimonio p ON m.patrimonio_id = p.id
        LEFT JOIN locais l ON p.local_id = l.id
        WHERE m.status = 'pendente'
        ORDER BY m.data_prevista
    """)
    for i, row in enumerate(cursor.fetchall(), 2):
        atrasado = (row["dias"] is not None and row["dias"] < 0)
        _estilo_linha(ws3, i, [
            row["codigo"], row["patrimonio"], row["local"], row["tipo"],
            row["descricao"], str(row["data_prevista"]),
            row["responsavel"], row["dias"]
        ], alerta=atrasado)
    _autofit(ws3)

    # ── Aba 4: Log de Auditoria ──────────────────────────────
    ws4 = wb.create_sheet("Auditoria")
    _estilo_header(ws4, 1, ["Data/Hora", "Usuário", "Ação", "Tabela", "ID Registro", "Descrição"])
    cursor.execute("""
        SELECT criado_em, usuario_login, acao, tabela, registro_id, descricao
        FROM auditoria ORDER BY criado_em DESC LIMIT 500
    """)
    for i, row in enumerate(cursor.fetchall(), 2):
        _estilo_linha(ws4, i, [
            str(row["criado_em"]), row["usuario_login"], row["acao"],
            row["tabela"], row["registro_id"], row["descricao"]
        ])
    _autofit(ws4)

    cursor.close()

    nome_arquivo = f"patrimonio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho = os.path.join(PASTA_RELATORIOS, nome_arquivo)
    wb.save(caminho)
    return caminho, "Relatório gerado com sucesso!"
